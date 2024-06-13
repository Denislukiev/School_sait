import pandas as pd
from googletrans import Translator
import requests
import datetime
import openpyxl
import codecs   
import sys
sys.stdout.reconfigure(encoding='utf-8')


try:
    today = datetime.date.today()
    print(today)

    # Формируем URL для загрузки файла
    url = f"http://food/{today}-sm.xlsx"
    response = requests.get(url)

    # Открываем файл для записи с указанием кодировки UTF-8
    with open("files.xlsx", "wb") as f:
        f.write(response.content)

    # Загружаем книгу Excel
    wb = openpyxl.load_workbook("files.xlsx")

    # Выбираем первый лист
    sheet = wb.active

    # Extract values for breakfast and lunch
    breakfast_values = [str(cell.value) for row in sheet.iter_rows(min_row=4, max_row=11, min_col=4, max_col=4) for cell in row if cell.value is not None]
    lunch_values = [str(cell.value) for row in sheet.iter_rows(min_row=12, max_row=31, min_col=4, max_col=4) for cell in row if cell.value is not None]

    # Translate breakfast values
    translator = Translator()
    translated_breakfast_values = []
    for value in breakfast_values:
        result = translator.translate(value, src='ru', dest='en')
        translated_breakfast_values.append(result.text)

    translator = Translator()
    translated_lunch_values = []
    for value in lunch_values:
        result = translator.translate(value, src='ru', dest='en')
        translated_lunch_values.append(result.text)
        # Generate HTML file
    with codecs.open('index.html', 'w', 'utf-8') as file:
        file.write('''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title id="menu-title">Меню</title>
    <style>

        /* Общие стили */
        body {
            font-family: Arial, sans-serif;
        }
        .menu-list {
        list-style: none;
        padding: 0;
        margin: 0;
        }
        .container-buttons
    {
        position: fixed;
            top: 100px; /* Располагаем вверху */
            left: 5px; /* Располагаем слева */
            z-index: 9999; /* Устанавливаем z-index, чтобы флаг отображался поверх других элементов */
    }
        .menu-list li {
        display: flex;
        align-items: baseline;
        }

        .menu-list li span {
        flex: 1;
        text-align: left;
        text-transform: capitalize;
        }
               
        .menu-list li::before {
        content: attr(data-number);
        display: inline-block;
        width: 2em;
        text-align: right;
        margin-right: 0.5em;
        font-weight: bold;
        font-size: 1.25em;
        }
        .flag-container {
            position: fixed;
            top: 50px; /* Располагаем вверху */
            left: 5px; /* Располагаем слева */
            z-index: 9999; /* Устанавливаем z-index, чтобы флаг отображался поверх других элементов */
        }
        .lungvich-button {
        display: inline-block; /* Отображение элемента как блочного элемента, но без переноса строки */
        padding: 15px 30px; /* Увеличиваем внутренние отступы */
        color: black; /* Цвет текста */
        text-decoration: none; /* Отмена подчёркивания ссылки */
        border-radius: 5px; /* Закругление углов элемента */
        border: none; /* Убрать границу элемента */
        background-image: linear-gradient(90deg, #ff0000, #ffee00, #f709ff); /* Фоновый градиент по горизонтали */
        background-size: 800% 800%; /* Размеры градиентного фона */
        animation: gradient 8s ease infinite; /* Анимация градиента */
        transition: background-color 0.5s ease; /* Добавлено свойство transition для плавного перехода */
        animation-delay: 1s; /* Добавлено значение задержки для синхронизации анимации */
    }

    @keyframes gradient {
        0% {
            background-position: 0% 50%; /* Начальная позиция фона градиента */
        }
        100% {
            background-position: 100% 50%; /* Конечная позиция фона градиента */
        }
    }

        .container {
            width: 80%;
            margin: auto;
            text-align: center;
        }
        .menu-button {
    display: inline-block; /* Отображение элемента как блочного элемента, но без переноса строки */
    margin: 30px; /* Увеличиваем внешние отступы */
    padding: 15px 30px; /* Увеличиваем внутренние отступы */
    color: black; /* Цвет текста */
    text-decoration: none; /* Отмена подчёркивания ссылки */
    border-radius: 5px; /* Закругление углов элемента */
    border: none; /* Убрать границу элемента */
    background-image: linear-gradient(90deg, #f8ef42, #0fd64f, #ffffff); /* Фоновый градиент по горизонтали */
    background-size: 800% 800%; /* Размеры градиентного фона */
    animation: gradient 8s ease infinite; /* Анимация градиента */
    transition: background-color 0.5s ease; /* Добавлено свойство transition для плавного перехода */
    animation-delay: 1s; /* Добавлено значение задержки для синхронизации анимации */
}

@keyframes gradient {
    0% {
        background-position: 0% 50%; /* Начальная позиция фона градиента */
    }
    100% {
        background-position: 100% 50%; /* Конечная позиция фона градиента */
    }
}


        /* Стили для кнопки при наведении курсора */
        .menu-button:hover {
            background-color: #ffffff; /* Цвет фона при наведении */
        }


        .menu-button:hover {
            background-color: #0cf552;
        }

        /* Стили для контента меню */
        .menu-content {
            display: none;
            margin-top: 20px;
            padding: 20px;
            border: 3px solid #000;
            border-radius: 5px;
        }
        .menu-content.active {
            display: block;
        }

        /* Добавленный контейнер для кнопок меню */
        .menu-buttons-container {
            display: flex;
            justify-content: center; /* Выравнивание кнопок по горизонтали по центру */
        }
        .top-right {
    position: fixed;
    top: 0;
    right: 0;
    padding: 10px;
    background-color: #f0f0f0;
}

    .rectangle {
    width: 96%; /* Ширина прямоугольника */
    height: 350px; /* Высота прямоугольника */
    background-color: #4CAF50; /* Цвет фона прямоугольника */
    border: 2px solid #000; /* Рамка вокруг прямоугольника */
    position: absolute; /* Абсолютное позиционирование */
    top: 540px; /* Расположение от верхнего края экрана */
    left: 20px; /* Расположение от левого края экрана */
}


    </style>
</head>
<body>
        <div class="photo-container">
            <img src="spons.png" alt="Your Photo" class="top-right"  width="350" height="150">
        </div>
        <div class="flag-container">
            <button id="changeTextButton" class="lungvich-button"> English
            </button>
        </div>
        <div class="datetime-container">
            <div id="date">{today.strftime("%d.%m.%Y")}</div>
            <div id="time">{datetime.datetime.now().strftime("%H:%M:%S")}</div>
        </div>
        <div class="container-buttons" id="123456789">

            </div>
        </div>
        <div class="container">
            <!-- Добавленный контейнер для кнопок меню -->
            <div class="menu-buttons-container">
                <!-- Кнопка "Завтрак" -->
                <button id="breakfast-button" class="menu-button">Завтрак</button>
                <!-- Кнопка "Обед" -->
                <button id="lunch-button" class="menu-button">Обед</button>
                <!-- Кнопка "полное меню" -->
                <button onclick="window.location.href='poln.html';" id="poln" class="menu-button">Полное меню</button>
            </div>

            <!-- Контент для завтрака -->
            <div id="breakfast-content" class="menu-content">
                <h2 id="breakfast">Завтрак</h2>
    <p id="p_breakfast">Список блюд для завтрака:</p>
                <ul>''')
        file.write('<ol class="menu-list" id="breakfast-list">' + ''.join(f'<li data-number="{i+1}.">{word}</li>' for i, word in enumerate(breakfast_values)) + '</ol>')
        file.write(f'''</ul>
        </div>

        <!-- Контент для обеда -->
        <div id="lunch-content" class="menu-content">
            <h2 id="lunch">Обед</h2>
            <p id="p_lunch">Список блюд для обеда:</p>
            <ul>''')
        file.write('<ol class="menu-list" id="lunch-list">' + ''.join(f'<li data-number="{i+1}.">{word}</li>' for i, word in enumerate(lunch_values)) + '</ol>')
        file.write('''</ul>
        ntById("lunch-button").textContent = "lunch";
        document.getElementById("poln").textContent = "Full menu";
        document.getElementById("breakfast").textContent = "breakfast";
        document.getElementById("lunch").textContent = "lunch";
        document.getElementById("p_breakfast").textContent = "List of breakfast dishes:";
        document.getElementById("p_lunch").textContent = "The list of dishes for lunch:";
        document.getElementById("breakfast-list").innerHTML = '  <li data-number="1.">LIQUID MILK OATH PORRIDGE "HERCULES" WITH BUTTER</li><li data-number="2.">COFFEE DRINK WITH MILK</li><li data-number="3.">LOAD SLICED</li><li data-number="4.">COOK PUDDING WITH CONDENSED MILK</li><li data-number="5.">APPLE FRESH</li> ';
        document.getElementById("lunch-list").innerHTML = ' <li data-number="1.">SALTED CUCUMBER</li><li data-number="2.">VEGETABLE SOUP WITH MEAT</li><li data-number="3.">LIVER STROGANOFF STYLE</li><li data-number="4.">MASHED POTATOES WITH CARROTS</li><li data-number="5.">COMPOTE FROM A MIXTURE OF DRIED FRUIT</li><li data-number="6.">LOAD SLICED</li><li data-number="7.">RYE-WHEAT BREAD</li><li data-number="8.">Cheesecake with cottage cheese</li> ';
        translationFlag = true;
        document.getElementById("changeTextButton").textContent = "Русский"; // изменяем текст кнопки
    } else {
        document.querySelectorAll("title").forEach(function(el) {
            el.textContent = "меню";
        });
        document.getElementById("breakfast-button").textContent = "завтрак";
        document.getElementById("lunch-button").textContent = "обед";
        document.getElementById("poln").textContent = "полное меню";
        document.getElementById("breakfast").textContent = "завтрак";
        document.getElementById("lunch").textContent = "обед";
        document.getElementById("p_breakfast").textContent = "Список блюд для завтрака:";
        document.getElementById("p_lunch").textContent = "Список блюд для обеда:";
        document.getElementById("breakfast-list").innerHTML = '  <li data-number="1.">КАША ЖИДКАЯ МОЛОЧНАЯ ОВСЯНАЯ "ГЕРКУЛЕС" С МАСЛОМ СЛИВОЧНЫМ</li><li data-number="2.">КОФЕЙНЫЙ НАПИТОК С МОЛОКОМ</li><li data-number="3.">БАТОН НАРЕЗНОЙ</li><li data-number="4.">ПУДИНГ ИЗ ТВОРОГА С МОЛОКОМ СГУЩЕННЫМ</li><li data-number="5.">ЯБЛОКО СВЕЖЕЕ</li> ';
        document.getElementById("lunch-list").innerHTML = ' <li data-number="1.">ОГУРЕЦ СОЛЕНЫЙ</li><li data-number="2.">СУП ИЗ ОВОЩЕЙ С МЯСОМ </li><li data-number="3.">ПЕЧЕНЬ ПО-СТРОГАНОВСКИ</li><li data-number="4.">ПЮРЕ КАРТОФЕЛЬНОЕ С МОРКОВЬЮ</li><li data-number="5.">КОМПОТ ИЗ СМЕСИ СУХОФРУКТОВ</li><li data-number="6.">БАТОН НАРЕЗНОЙ</li><li data-number="7.">ХЛЕБ РЖАНО-ПШЕНИЧНЫЙ</li><li data-number="8.">ВАТРУШКА С ТВОРОГОМ</li> ';
 document.getElementById("lunch-list").innerHTML = ' <li data-number="1.">ОГУРЕЦ СОЛЕНЫЙ</li><li data-number="2.">СУП ИЗ ОВОЩЕЙ С МЯСОМ </li><li data-number="3.">ПЕЧЕНЬ ПО-СТРОГАНОВСКИ</li><li data-number="4.">ПЮРЕ КАРТОФЕЛЬНОЕ С МОРКОВЬЮ</li><li data-number="5.">КОМПОТ ИЗ СМЕСИ СУХОФРУКТОВ</li><li data-number="6.">БАТОН НАРЕЗНОЙ</li><li data-number="7.">ХЛЕБ РЖАНО-ПШЕНИЧНЫЙ</li><li data-number="8.">ВАТРУШКА С ТВОРОГОМ</li> ';
            translationFlag = false;
        document.getElementById("changeTextButton").textContent = "English"; // изменяем текст кнопки
    }
});

    setInterval(() => {
        const timeSinceLastClick = Date.now() - lastClickTime;
        if (timeSinceLastClick > REFRESH_INTERVAL) {
            window.location.reload();
        } else {
            window.scrollTo(0, 0);
        }
        }, 1000); 
        function updateDateTime() {
            var now = new Date();
            var options = { day: '2-digit', month: '2-digit', year: 'numeric' };
            var dateElement = document.getElementById('date');
            var timeElement = document.getElementById('time');

            // Форматируем дату в формат "дд-мм-гггг"
            var formattedDate = now.toLocaleDateString('ru-RU', options);
            dateElement.textContent = formattedDate;

            // Получаем точное время по Московскому времени
            var mskTime = now.toLocaleTimeString('ru-RU', { timeZone: 'Europe/Moscow', hour12: false });
            timeElement.textContent = mskTime;
        }
        // Обновляем дату и время каждую секунду
        setInterval(updateDateTime, 1000);

        // Вызываем функцию один раз для отображения начального значения
        updateDateTime();

        // Функция для переключения контента и кнопок
function toggleContent(showContentId, hideContentId, showButtonId, hideButtonId) {
    // Скрываем контент, который нужно скрыть
    document.getElementById(hideContentId).classList.remove('active');
    // Отображаем контент, который нужно показать
    document.getElementById(showContentId).classList.add('active');
    // Скрываем кнопку, которую нужно скрыть
    document.getElementById(hideButtonId).style.display = 'none';
    // Отображаем кнопку, которую нужно показать
    document.getElementById(showButtonId).style.display = 'inline-block'; // изменено на 'inline-block'

    // Скрываем элемент с классом 'bottom-right'
    document.querySelector('.bottom-right').style.display = 'none';
}

// Вешаем обработчик события на кнопку "Завтрак"
document.getElementById('breakfast-button').addEventListener('click', function() {
    toggleContent('breakfast-content', 'lunch-content', 'lunch-button', 'breakfast-button');
});

// Вешаем обработчик события на кнопку "Обед"
document.getElementById('lunch-button').addEventListener('click', function() {
    toggleContent('lunch-content', 'breakfast-content', 'breakfast-button', 'lunch-button');
});

// Функция, которая будет вызываться при клике на кнопку
function handleClick(event) {
  // Получаем ID кнопки, которая была нажата
  const buttonId = event.target.id;

  // Проверяем, является ли кнопка одной из разрешенных
  if (allowedButtonIds.includes(buttonId)) {
    // Удаляем кнопку
    event.target.style.display = 'none';

    // Перемещаем другие кнопки в заданное место
    const otherButtons = buttons.filter(button => button.id!== buttonId && allowedButtonIds.includes(button.id));
    otherButtons.forEach(button => {
      targetElement.appendChild(button);
    });
  }
}

// Добавляем обработчик события клика на каждую кнопку
buttons.forEach(button => {
  button.addEventListener('click', handleClick);
});

        (function() {{
  var ws = new WebSocket('ws://' + window.location.host +
             '/jb-server-page?reloadMode=RELOAD_ON_SAVE&'+
             'referrer=' + encodeURIComponent(window.location.pathname));
  ws.onmessage = function (msg) {{
      if (msg.data === 'reload') {{
          window.location.reload();
      }}
      if (msg.data.startsWith('update-css ')) {{
          var messageId = msg.data.substring(11);
          var links = document.getElementsByTagName('link');
          for (var i = 0; i < links.length; i++) {{
              var link = links[i];
              if (link.rel !== 'stylesheet') continue;
              var clonedLink = link.cloneNode(true);
              var newHref = link.href.replace(/(&|\?)jbUpdateLinksId=\d+/, "$1jbUpdateLinksId=" + messageId);
              if (newHref !== link.href) {{
                clonedLink.href = newHref;
              }}
              else {{
                var indexOfQuest = newHref.indexOf('?');
                if (indexOfQuest >= 0) {{
                  // to support ?foo#hash
                  clonedLink.href = newHref.substring(0, indexOfQuest + 1) + 'jbUpdateLinksId=' + messageId + '&' +
                                    newHref.substring(indexOfQuest + 1);
                }}
                else {{
                  clonedLink.href += '?' + 'jbUpdateLinksId=' + messageId;
                }}
              }}
              link.replaceWith(clonedLink);
          }}
      }}
  }};
}})();
</script>
</body>
</html>''')

except Exception as e: 
    error_message = '<img src="spons.png" style="width:100%">' 
    with open('index.html', 'w') as file: 
        file.write(error_message)
